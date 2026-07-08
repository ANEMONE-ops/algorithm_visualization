"""
算法可视化系统 - 日志导出模块 (导出为 .xlsx 格式)
"""

import io
import json
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..database import get_db
from ..models import User, ExecutionLog
from .auth import get_current_user

router = APIRouter(prefix="/api/export", tags=["日志导出"])

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

ALGO_NAMES = {
    "bubble_sort": "冒泡排序",
    "quick_sort": "快速排序",
    "mst": "最小生成树(Prim)",
    "huffman": "哈夫曼树",
    "hanoi": "汉诺塔",
    "graph_coloring": "图着色(回溯)",
}


class SaveLogRequest(BaseModel):
    algorithm_type: str
    input_data: str
    is_test_case: int = 0
    test_case_index: Optional[int] = None
    total_steps: int
    execution_time_ms: Optional[int] = None


# ---------------------------------------------------------------------------
# Excel 生成
# ---------------------------------------------------------------------------

def make_xlsx(logs: list, title: str = "算法执行日志") -> io.BytesIO:
    """将日志列表生成为 .xlsx 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "执行日志"

    # 样式
    header_font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4A90D9", end_color="4A90D9", fill_type="solid",
    )
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Microsoft YaHei", size=11)
    cell_align = Alignment(
        horizontal="center", vertical="center", wrap_text=True,
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 标题行
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(
        name="Microsoft YaHei", size=16, bold=True, color="333333",
    )
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 导出时间
    ws.merge_cells("A2:H2")
    time_cell = ws.cell(
        row=2, column=1,
        value=f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}    共 {len(logs)} 条记录",
    )
    time_cell.font = Font(name="Microsoft YaHei", size=10, color="888888")
    time_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 24

    # 表头
    headers = [
        "序号", "算法名称", "输入数据", "是否测试用例",
        "测试用例编号", "总步数", "执行耗时(ms)", "执行时间",
    ]
    col_widths = [8, 22, 45, 14, 14, 10, 14, 22]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[4].height = 28

    # 数据行
    for row_idx, log in enumerate(logs, 1):
        row_num = row_idx + 4
        algo_name = ALGO_NAMES.get(
            log.get("algorithm_type", ""),
            log.get("algorithm_type", ""),
        )

        input_str = log.get("input_data", "")
        try:
            parsed = json.loads(input_str)
            input_str = json.dumps(parsed, ensure_ascii=False)
        except Exception:
            pass
        if len(input_str) > 100:
            input_str = input_str[:100] + "..."

        values = [
            row_idx,
            algo_name,
            input_str,
            "是" if log.get("is_test_case") else "否",
            log.get("test_case_index") or "-",
            log.get("total_steps", 0),
            log.get("execution_time_ms") or "-",
            log.get("created_at", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

        # 交替行颜色
        if row_idx % 2 == 0:
            row_fill = PatternFill(
                start_color="F5F8FC", end_color="F5F8FC", fill_type="solid",
            )
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_idx).fill = row_fill

        ws.row_dimensions[row_num].height = 24

    ws.freeze_panes = "A5"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# 端点
# ---------------------------------------------------------------------------

@router.post("/save-log")
def save_execution_log(
    req: SaveLogRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """保存算法执行日志"""
    log = ExecutionLog(
        user_id=current_user.id,
        algorithm_type=req.algorithm_type,
        input_data=req.input_data,
        is_test_case=req.is_test_case,
        test_case_index=req.test_case_index,
        total_steps=req.total_steps,
        execution_time_ms=req.execution_time_ms,
    )
    db.add(log)
    db.commit()
    return {"message": "日志保存成功", "log_id": log.id}


@router.get("/logs")
def get_user_logs(
    algorithm_type: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=200),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取用户的执行日志"""
    query = db.query(ExecutionLog).filter(
        ExecutionLog.user_id == current_user.id,
    )
    if algorithm_type:
        query = query.filter(ExecutionLog.algorithm_type == algorithm_type)
    logs = query.order_by(ExecutionLog.created_at.desc()).limit(limit).all()

    return {
        "logs": [
            {
                "id": log.id,
                "algorithm_type": log.algorithm_type,
                "input_data": log.input_data,
                "is_test_case": log.is_test_case,
                "test_case_index": log.test_case_index,
                "total_steps": log.total_steps,
                "execution_time_ms": log.execution_time_ms,
                "created_at": (
                    log.created_at.isoformat() if log.created_at else None
                ),
            }
            for log in logs
        ],
    }


@router.get("/export/{log_id}")
def export_single_log(
    log_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出单条日志为 .xlsx"""
    log = db.query(ExecutionLog).filter(
        ExecutionLog.id == log_id,
        ExecutionLog.user_id == current_user.id,
    ).first()

    if not log:
        raise HTTPException(status_code=404, detail="日志不存在")

    log_data = [{
        "algorithm_type": log.algorithm_type,
        "input_data": log.input_data,
        "is_test_case": log.is_test_case,
        "test_case_index": log.test_case_index,
        "total_steps": log.total_steps,
        "execution_time_ms": log.execution_time_ms,
        "created_at": log.created_at.isoformat() if log.created_at else "",
    }]

    xlsx = make_xlsx(log_data, title=f"算法执行日志 #{log_id}")
    filename = f"algo_log_{log_id}.xlsx"

    return StreamingResponse(
        xlsx,
        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export-all")
def export_all_logs(
    algorithm_type: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """批量导出全部日志为 .xlsx"""
    query = db.query(ExecutionLog).filter(
        ExecutionLog.user_id == current_user.id,
    )
    if algorithm_type:
        query = query.filter(ExecutionLog.algorithm_type == algorithm_type)
    logs = query.order_by(ExecutionLog.created_at.desc()).all()

    log_data = [
        {
            "algorithm_type": log.algorithm_type,
            "input_data": log.input_data,
            "is_test_case": log.is_test_case,
            "test_case_index": log.test_case_index,
            "total_steps": log.total_steps,
            "execution_time_ms": log.execution_time_ms,
            "created_at": log.created_at.isoformat() if log.created_at else "",
        }
        for log in logs
    ]

    xlsx = make_xlsx(log_data, title="算法执行日志汇总")
    filename = f"algo_logs_all_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        xlsx,
        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
